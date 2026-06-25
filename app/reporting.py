"""Excel report generation for shortlisted candidates (openpyxl, in-memory)."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Columns per the recruitment spec (Ranking first).
_COLUMNS = [
    "Rank", "Candidate Name", "Email", "Match Score", "Recommendation",
    "Relevant Skills", "Experience", "Education", "Projects", "Certifications",
    "Selection Reasons",
]
_WIDTHS = [6, 22, 26, 11, 16, 38, 14, 30, 38, 28, 50]


def _join(v) -> str:
    if isinstance(v, (list, tuple)):
        return "; ".join(str(x) for x in v if x not in (None, ""))
    return "" if v in (None, "") else str(v)


def _experience(r: dict) -> str:
    yrs = r.get("years_experience_estimate")
    head = f"{yrs} yrs" if yrs not in (None, "") else ""
    detail = _join(r.get("experience"))
    return f"{head} — {detail}" if head and detail else (head or detail)


def _selection_reason(r: dict) -> str:
    parts = []
    if r.get("summary"):
        parts.append(r["summary"])
    if r.get("strengths"):
        parts.append("Strengths: " + _join(r.get("strengths")))
    return "  ".join(parts)


def shortlisted_rows(run: dict) -> list[dict]:
    """Shortlisted candidates, best first (mirrors the API's ranking order)."""
    rows = [r for r in run.get("results", []) if r.get("shortlisted")]
    rows.sort(
        key=lambda r: (
            r.get("overall_score") is not None,
            r.get("overall_score") if r.get("overall_score") is not None else 0,
            r.get("similarity", 0),
        ),
        reverse=True,
    )
    return rows


def shortlist_xlsx(run: dict) -> bytes:
    """Build an .xlsx workbook (bytes) of the run's shortlisted candidates."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shortlist"

    ws.append(_COLUMNS)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for rank, r in enumerate(shortlisted_rows(run), start=1):
        ws.append([
            rank,
            r.get("candidate_name") or r.get("filename", ""),
            r.get("candidate_email") or "",
            r.get("overall_score") if r.get("overall_score") is not None else "",
            r.get("recommendation") or "",
            _join(r.get("required_skills_matched")),
            _experience(r),
            _join(r.get("education")),
            _join(r.get("projects")),
            _join(r.get("certifications")),
            _selection_reason(r),
        ])

    for i, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
