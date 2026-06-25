"""Excel shortlist report — builds a valid workbook with the right columns,
ranking order, and the extracted CV fields."""

import io

from openpyxl import load_workbook

from app import reporting

RUN = {
    "id": "abc123",
    "results": [
        {"filename": "a.pdf", "candidate_name": "Alice", "candidate_email": "alice@x.com",
         "shortlisted": True, "overall_score": 88, "similarity": 0.9,
         "recommendation": "Strong Match", "required_skills_matched": ["Python", "SQL"],
         "years_experience_estimate": 6, "experience": ["Engineer at X (2019-2024)"],
         "education": ["B.Tech, IIT"], "projects": ["Billing system"],
         "certifications": ["PMP"], "strengths": ["Leadership"], "summary": "Strong fit."},
        {"filename": "b.pdf", "candidate_name": "Bob", "candidate_email": "bob@x.com",
         "shortlisted": True, "overall_score": 72, "similarity": 0.8,
         "recommendation": "Good Match", "required_skills_matched": ["SQL"],
         "education": [], "projects": [], "certifications": [], "strengths": [], "summary": "OK."},
        {"filename": "c.pdf", "candidate_name": "Carol", "shortlisted": False,
         "overall_score": None, "similarity": 0.3, "recommendation": "Not shortlisted"},
    ],
}


def test_xlsx_is_valid_and_ranked():
    data = reporting.shortlist_xlsx(RUN)
    assert data[:2] == b"PK"  # xlsx is a zip
    ws = load_workbook(io.BytesIO(data)).active
    assert ws["A1"].value == "Rank"
    assert ws["B1"].value == "Candidate Name"
    # only the 2 shortlisted, best score first
    assert ws["B2"].value == "Alice" and ws["A2"].value == 1
    assert ws["B3"].value == "Bob" and ws["A3"].value == 2
    assert ws.max_row == 3  # header + 2
    # email + extracted fields carried through
    assert ws["C2"].value == "alice@x.com"
    assert "Python" in ws["F2"].value and "SQL" in ws["F2"].value
    assert "IIT" in ws["H2"].value


def test_xlsx_empty_run_has_header_only():
    ws = load_workbook(io.BytesIO(reporting.shortlist_xlsx({"results": []}))).active
    assert ws["A1"].value == "Rank" and ws.max_row == 1
